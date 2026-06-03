"""Layer 1: XLSX ACTION entity definitions for Document Factory Engine.

Distilled from Claude Skills: skills/xlsx/SKILL.md
"""

ACTIONS_XLSX = [
    {
        "key": "xlsx_create_action",
        "payload": {
            "name": "doc-create-xlsx",
            "display_name": "XLSX Creator (openpyxl)",
            "description": "Creates new Excel workbooks with data, formulas, formatting, and charts using openpyxl.",
            "goal": "Generate a professional Excel workbook with proper formulas (never hardcoded values), formatting, and multiple sheets.",
            "type": "ACTION", "version": "1.0.0", "status": "ACTIVE",
            "tags": ["doc-factory", "xlsx", "creation", "openpyxl"],
            "identity": {
                "system_prompt": (
                    "You are an Excel specialist. Create workbooks using openpyxl.\n\n"
                    "## CRITICAL: USE FORMULAS, NOT HARDCODED VALUES\n"
                    "```python\n"
                    "# ❌ WRONG: total = df['Sales'].sum(); sheet['B10'] = total\n"
                    "# ✅ CORRECT: sheet['B10'] = '=SUM(B2:B9)'\n"
                    "# ❌ WRONG: growth = (val2-val1)/val1; sheet['C5'] = growth\n"
                    "# ✅ CORRECT: sheet['C5'] = '=(C4-C2)/C2'\n"
                    "```\n\n"
                    "## CREATION\n"
                    "```python\n"
                    "from openpyxl import Workbook\n"
                    "from openpyxl.styles import Font, PatternFill, Alignment\n\n"
                    "wb = Workbook()\n"
                    "sheet = wb.active\n"
                    "sheet['A1'] = 'Header'\n"
                    "sheet['A1'].font = Font(bold=True, color='FF0000')\n"
                    "sheet['A1'].fill = PatternFill('solid', start_color='FFFF00')\n"
                    "sheet['A1'].alignment = Alignment(horizontal='center')\n"
                    "sheet.column_dimensions['A'].width = 20\n"
                    "sheet['B2'] = '=SUM(A1:A10)'  # USE FORMULAS!\n"
                    "wb.save('output.xlsx')\n"
                    "```\n\n"
                    "## POST-CREATION: RECALCULATE (MANDATORY)\n"
                    "```bash\n"
                    "python scripts/xlsx/recalc.py output.xlsx\n"
                    "```\n"
                    "This recalculates all formulas and scans for errors (#REF!, #DIV/0!, etc.).\n\n"
                    "## PROFESSIONAL FONT\n"
                    "Use Arial or Times New Roman consistently.\n\n"
                    "## ZERO FORMULA ERRORS\n"
                    "Every workbook MUST be delivered with ZERO formula errors.\n\n"
                    "## CODE STYLE\n"
                    "Write minimal, concise Python. Avoid verbose variable names and unnecessary comments."
                ),
                "behavioral_constraints": [
                    "ALWAYS use Excel formulas instead of calculating values in Python and hardcoding them",
                    "Every workbook must be delivered with ZERO formula errors (#REF!, #DIV/0!, #VALUE!, #N/A, #NAME?)",
                    "ALWAYS run scripts/xlsx/recalc.py after creating workbooks with formulas",
                    "Use consistent professional font (Arial or Times New Roman)",
                    "Write minimal, concise Python code without unnecessary comments"
                ]
            },
            "hierarchy": {"is_atomic": True, "composition_depth": 0, "children": []},
            "logic_gate": {
                "reasoning_config": {"task_type": "text_generation", "temperature": 0.3, "reasoning_mode": "CHAIN_OF_THOUGHT"},
                "retry_policy": {"max_retries": 2, "backoff_strategy": "EXPONENTIAL"},
                "context_policy": {"type": "FULL", "summarize_threshold": 20000}
            },
            "planning": {"static_plan": {"enabled": True, "steps": [{"step_id": "step_1", "order": 1, "name": "Generate XLSX", "type": "ACTION", "target": {"tool_id": "sandbox_code", "prompt_template": "{{input}}"}, "required": True}]}},
            "capabilities": {"tools": [{"tool_id": "sandbox_code"}], "memory": {"enabled": True, "mode": "CORTEX"}},
            "governance": {"timeout_ms": 300000, "max_cost_usd": 0.50},
            "io_contract": {
                "input_schema": {"type": "object", "properties": {"input": {"type": "string", "description": "Task input for this action"}}, "required": ["input"]},
                "output_schema": {"type": "object", "properties": {"result": {"type": "string", "description": "Action execution result"}}}
            }
        }
    },
    {
        "key": "xlsx_edit_action",
        "payload": {
            "name": "doc-edit-xlsx",
            "display_name": "XLSX Editor (openpyxl)",
            "description": "Edits existing Excel files — modifies cells, adds sheets, inserts/deletes rows/columns while preserving formatting.",
            "goal": "Apply precise modifications to existing Excel workbooks while preserving all existing formatting and formulas.",
            "type": "ACTION", "version": "1.0.0", "status": "ACTIVE",
            "tags": ["doc-factory", "xlsx", "editing", "openpyxl"],
            "identity": {
                "system_prompt": (
                    "You edit existing Excel files using openpyxl.\n\n"
                    "```python\n"
                    "from openpyxl import load_workbook\n"
                    "wb = load_workbook('existing.xlsx')\n"
                    "sheet = wb.active  # or wb['SheetName']\n\n"
                    "# Modify cells\n"
                    "sheet['A1'] = 'New Value'\n"
                    "sheet.insert_rows(2)\n"
                    "sheet.delete_cols(3)\n\n"
                    "# Add new sheet\n"
                    "new_sheet = wb.create_sheet('NewSheet')\n"
                    "wb.save('modified.xlsx')\n"
                    "```\n\n"
                    "## PRESERVE EXISTING TEMPLATES\n"
                    "Study and EXACTLY match existing format, style, and conventions.\n"
                    "Never impose standardized formatting on files with established patterns.\n"
                    "Existing template conventions ALWAYS override default guidelines.\n\n"
                    "## WARNING: data_only=True\n"
                    "load_workbook('file.xlsx', data_only=True) reads calculated values but if saved,\n"
                    "formulas are replaced with values and PERMANENTLY LOST."
                ),
                "behavioral_constraints": [
                    "Study and exactly match existing format/style/conventions when modifying files",
                    "Never impose standardized formatting on files with established patterns",
                    "NEVER open with data_only=True if you plan to save — it permanently destroys formulas",
                    "Run recalc.py after any formula modifications"
                ]
            },
            "hierarchy": {"is_atomic": True, "composition_depth": 0, "children": []},
            "logic_gate": {"reasoning_config": {"task_type": "text_generation", "temperature": 0.2, "reasoning_mode": "CHAIN_OF_THOUGHT"}, "context_policy": {"type": "FULL", "summarize_threshold": 20000}},
            "planning": {"static_plan": {"enabled": True, "steps": [{"step_id": "step_1", "order": 1, "name": "Edit XLSX", "type": "ACTION", "target": {"tool_id": "sandbox_code", "prompt_template": "{{input}}"}, "required": True}]}},
            "capabilities": {"tools": [{"tool_id": "sandbox_code"}], "memory": {"enabled": True, "mode": "CORTEX"}},
            "governance": {"timeout_ms": 300000, "max_cost_usd": 0.50},
            "io_contract": {
                "input_schema": {"type": "object", "properties": {"input": {"type": "string", "description": "Task input for this action"}}, "required": ["input"]},
                "output_schema": {"type": "object", "properties": {"result": {"type": "string", "description": "Action execution result"}}}
            }
        }
    },
    {
        "key": "xlsx_analyze_data_action",
        "payload": {
            "name": "doc-analyze-xlsx-data",
            "display_name": "XLSX Data Analyzer (pandas)",
            "description": "Analyzes Excel data using pandas — statistics, visualization, and bulk operations.",
            "goal": "Extract insights from Excel data using pandas for analysis, aggregation, and visualization.",
            "type": "ACTION", "version": "1.0.0", "status": "ACTIVE",
            "tags": ["doc-factory", "xlsx", "analysis", "pandas"],
            "identity": {
                "system_prompt": (
                    "You analyze Excel data using pandas.\n\n"
                    "```python\n"
                    "import pandas as pd\n"
                    "df = pd.read_excel('file.xlsx')  # first sheet\n"
                    "all_sheets = pd.read_excel('file.xlsx', sheet_name=None)  # all sheets as dict\n"
                    "df.head(); df.info(); df.describe()\n"
                    "df.to_excel('output.xlsx', index=False)\n"
                    "```\n\n"
                    "## TIPS\n"
                    "- Specify dtypes: pd.read_excel('f.xlsx', dtype={'id': str})\n"
                    "- Read specific columns: usecols=['A', 'C', 'E']\n"
                    "- Handle dates: parse_dates=['date_column']\n"
                    "- For large files: read_only=True or specific columns"
                ),
                "behavioral_constraints": [
                    "Specify data types to avoid inference issues",
                    "Handle NaN values with pd.notna()",
                    "Report data shape, dtypes, and basic statistics"
                ]
            },
            "hierarchy": {"is_atomic": True, "composition_depth": 0, "children": []},
            "logic_gate": {"reasoning_config": {"task_type": "text_generation", "temperature": 0.2, "reasoning_mode": "CHAIN_OF_THOUGHT"}, "context_policy": {"type": "FULL", "summarize_threshold": 15000}},
            "planning": {"static_plan": {"enabled": True, "steps": [{"step_id": "step_1", "order": 1, "name": "Analyze Data", "type": "ACTION", "target": {"tool_id": "sandbox_code", "prompt_template": "{{input}}"}, "required": True}]}},
            "capabilities": {"tools": [{"tool_id": "sandbox_code"}], "memory": {"enabled": True, "mode": "CORTEX"}},
            "governance": {"timeout_ms": 300000, "max_cost_usd": 0.50},
            "io_contract": {
                "input_schema": {"type": "object", "properties": {"input": {"type": "string", "description": "Task input for this action"}}, "required": ["input"]},
                "output_schema": {"type": "object", "properties": {"result": {"type": "string", "description": "Action execution result"}}}
            }
        }
    },
    {
        "key": "xlsx_recalc_action",
        "payload": {
            "name": "doc-recalc-xlsx",
            "display_name": "XLSX Formula Recalculator",
            "description": "Recalculates all formulas in an Excel file using LibreOffice and scans for errors.",
            "goal": "Recalculate all formulas and identify any formula errors for correction.",
            "type": "ACTION", "version": "1.0.0", "status": "ACTIVE",
            "tags": ["doc-factory", "xlsx", "formulas", "recalculation"],
            "identity": {
                "system_prompt": (
                    "You recalculate Excel formulas and verify correctness.\n\n"
                    "```bash\npython scripts/xlsx/recalc.py output.xlsx [timeout_seconds]\n```\n\n"
                    "## OUTPUT FORMAT (JSON)\n"
                    "```json\n"
                    "{ \"status\": \"success\", \"total_errors\": 0, \"total_formulas\": 42 }\n"
                    "// or\n"
                    "{ \"status\": \"errors_found\", \"total_errors\": 2,\n"
                    "  \"error_summary\": { \"#REF!\": { \"count\": 2, \"locations\": [\"Sheet1!B5\"] } } }\n"
                    "```\n\n"
                    "## COMMON ERRORS\n"
                    "- #REF!: Invalid cell references\n"
                    "- #DIV/0!: Division by zero\n"
                    "- #VALUE!: Wrong data type in formula\n"
                    "- #NAME?: Unrecognized formula name"
                ),
                "behavioral_constraints": [
                    "If errors_found, report exact error types and locations",
                    "Common fixes: verify cell references, check denominators, verify cross-sheet format (Sheet1!A1)"
                ]
            },
            "hierarchy": {"is_atomic": True, "composition_depth": 0, "children": []},
            "logic_gate": {"reasoning_config": {"task_type": "text_generation", "temperature": 0.1, "reasoning_mode": "REACT"}},
            "planning": {"static_plan": {"enabled": True, "steps": [{"step_id": "step_1", "order": 1, "name": "Recalculate Formulas", "type": "ACTION", "target": {"tool_id": "terminal", "prompt_template": "{{input}}"}, "required": True}]}},
            "capabilities": {"tools": [{"tool_id": "terminal"}], "memory": {"enabled": True, "mode": "CORTEX"}},
            "governance": {"timeout_ms": 300000, "max_cost_usd": 0.50},
            "io_contract": {
                "input_schema": {"type": "object", "properties": {"input": {"type": "string", "description": "Task input for this action"}}, "required": ["input"]},
                "output_schema": {"type": "object", "properties": {"result": {"type": "string", "description": "Action execution result"}}}
            }
        }
    },
    {
        "key": "xlsx_verify_errors_action",
        "payload": {
            "name": "doc-verify-xlsx-errors",
            "display_name": "XLSX Error Verifier & Fixer",
            "description": "Analyzes recalculation error output and applies fixes to Excel formulas.",
            "goal": "Fix all formula errors identified by recalculation until zero errors remain.",
            "type": "ACTION", "version": "1.0.0", "status": "ACTIVE",
            "tags": ["doc-factory", "xlsx", "error-fixing", "formulas"],
            "identity": {
                "system_prompt": (
                    "You fix Excel formula errors identified by recalc.py.\n\n"
                    "## VERIFICATION CHECKLIST\n"
                    "- Test 2-3 sample references before building full model\n"
                    "- Column mapping: column 64 = BL, not BK\n"
                    "- Row offset: Excel rows are 1-indexed (DataFrame row 5 = Excel row 6)\n"
                    "- NaN handling: check with pd.notna()\n"
                    "- Division by zero: check denominators\n"
                    "- Cross-sheet references: use format Sheet1!A1\n"
                    "- Start small: test on 2-3 cells before applying broadly\n\n"
                    "## FIX STRATEGY\n"
                    "1. Load workbook (NOT with data_only=True)\n"
                    "2. Navigate to error cells\n"
                    "3. Diagnose: check referenced cells exist and contain expected data\n"
                    "4. Fix formula\n"
                    "5. Save and re-run recalc.py\n"
                    "6. Repeat until zero errors"
                ),
                "behavioral_constraints": [
                    "Never open with data_only=True when fixing — destroys formulas permanently",
                    "Test fixes on 2-3 cells before applying broadly",
                    "Repeat fix-and-verify until zero errors remain"
                ]
            },
            "hierarchy": {"is_atomic": True, "composition_depth": 0, "children": []},
            "logic_gate": {"reasoning_config": {"task_type": "text_generation", "temperature": 0.2, "reasoning_mode": "CHAIN_OF_THOUGHT"}, "context_policy": {"type": "FULL", "summarize_threshold": 15000}},
            "planning": {"static_plan": {"enabled": True, "steps": [{"step_id": "step_1", "order": 1, "name": "Fix Formula Errors", "type": "ACTION", "target": {"tool_id": "sandbox_code", "prompt_template": "{{input}}"}, "required": True}]}},
            "capabilities": {"tools": [{"tool_id": "sandbox_code"}], "memory": {"enabled": True, "mode": "CORTEX"}},
            "governance": {"timeout_ms": 300000, "max_cost_usd": 0.50},
            "io_contract": {
                "input_schema": {"type": "object", "properties": {"input": {"type": "string", "description": "Task input for this action"}}, "required": ["input"]},
                "output_schema": {"type": "object", "properties": {"result": {"type": "string", "description": "Action execution result"}}}
            }
        }
    },
    {
        "key": "xlsx_financial_format_action",
        "payload": {
            "name": "doc-xlsx-financial-format",
            "display_name": "XLSX Financial Formatter",
            "description": "Applies industry-standard financial model formatting: color coding, number formats, assumption placement, and documentation.",
            "goal": "Transform a workbook into a professionally formatted financial model following IB/consulting standards.",
            "type": "ACTION", "version": "1.0.0", "status": "ACTIVE",
            "tags": ["doc-factory", "xlsx", "financial", "formatting"],
            "identity": {
                "system_prompt": (
                    "You apply financial model formatting standards.\n\n"
                    "## COLOR CODING (Industry Standard)\n"
                    "- Blue text (0,0,255): Hardcoded inputs, scenario numbers\n"
                    "- Black text (0,0,0): ALL formulas and calculations\n"
                    "- Green text (0,128,0): Links from other worksheets in same workbook\n"
                    "- Red text (255,0,0): External links to other files\n"
                    "- Yellow background (255,255,0): Key assumptions needing attention\n\n"
                    "## NUMBER FORMATTING\n"
                    "- Years: Format as text strings ('2024' not '2,024')\n"
                    "- Currency: $#,##0 format; ALWAYS specify units in headers ('Revenue ($mm)')\n"
                    "- Zeros: Display as '-' using format '$#,##0;($#,##0);-'\n"
                    "- Percentages: 0.0% (one decimal)\n"
                    "- Multiples: 0.0x for valuation multiples\n"
                    "- Negatives: Parentheses (123) not minus -123\n\n"
                    "## FORMULA RULES\n"
                    "- Place ALL assumptions in separate cells, never hardcode in formulas\n"
                    "- Use =B5*(1+$B$6) instead of =B5*1.05\n"
                    "- Verify no circular references\n"
                    "- Consistent formulas across all projection periods\n\n"
                    "## DOCUMENTATION FOR HARDCODES\n"
                    "Format: 'Source: [System/Document], [Date], [Reference], [URL]'\n"
                    "Example: 'Source: Company 10-K, FY2024, Page 45, Revenue Note, [SEC EDGAR URL]'"
                ),
                "behavioral_constraints": [
                    "Blue text for hardcoded inputs, black for formulas — never mix",
                    "Always specify units in headers (Revenue ($mm))",
                    "Display zeros as '-' using proper number format",
                    "Use parentheses for negatives, never minus sign",
                    "Place all assumptions in separate cells — never hardcode in formulas",
                    "Document sources for all hardcoded values"
                ]
            },
            "hierarchy": {"is_atomic": True, "composition_depth": 0, "children": []},
            "logic_gate": {"reasoning_config": {"task_type": "text_generation", "temperature": 0.2, "reasoning_mode": "CHAIN_OF_THOUGHT"}, "context_policy": {"type": "FULL", "summarize_threshold": 20000}},
            "planning": {"static_plan": {"enabled": True, "steps": [{"step_id": "step_1", "order": 1, "name": "Apply Financial Formatting", "type": "ACTION", "target": {"tool_id": "sandbox_code", "prompt_template": "{{input}}"}, "required": True}]}},
            "capabilities": {"tools": [{"tool_id": "sandbox_code"}], "memory": {"enabled": True, "mode": "CORTEX"}},
            "governance": {"timeout_ms": 300000, "max_cost_usd": 0.50},
            "io_contract": {
                "input_schema": {"type": "object", "properties": {"input": {"type": "string", "description": "Task input for this action"}}, "required": ["input"]},
                "output_schema": {"type": "object", "properties": {"result": {"type": "string", "description": "Action execution result"}}}
            }
        }
    },
]
