"""
Campaign API Router for frontend integration.

Provides endpoints for:
- Creating campaigns
- Uploading contact lists
- Managing campaign execution
- Retrieving status and metrics
- Downloading campaign reports as Excel
"""
import io
import logging
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from uuid import UUID
from typing import List, Dict, Any, Optional
from datetime import datetime

from src.common.database import get_db
from src.auth.dependencies import get_current_user
from src.auth.models import User
from src.ai.campaign_service import CampaignService
from src.ai.models import HierarchicalEntity

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/campaigns", tags=["Campaigns"])


# Pydantic models for request/response
class CampaignCreate(BaseModel):
    """Campaign creation request."""
    agent_id: UUID
    name: str
    description: Optional[str] = None
    contact_list: List[Dict[str, Any]]
    provider: str = "twilio"
    call_script_template: Optional[str] = None
    scheduled_start: Optional[datetime] = None
    scheduled_end: Optional[datetime] = None
    max_concurrent_calls: int = 5
    max_calls_per_hour: Optional[int] = None
    metadata: Optional[Dict[str, Any]] = None


class CampaignResponse(BaseModel):
    """Campaign response model."""
    id: str
    name: str
    description: Optional[str]
    status: str
    total_contacts: int
    calls_initiated: int
    calls_completed: int
    calls_failed: int
    created_at: str


@router.post("")
async def create_campaign(
    campaign_data: CampaignCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Create a new voice calling campaign.

    Resolves the campaign's company_id from the selected agent's company,
    NOT the current user's company. This is critical for multi-tenant setups
    where an app/partner admin creates campaigns on behalf of a tenant —
    the campaign must be bound to the tenant's company so that the correct
    phone number (assigned to the tenant) is used for outbound calls.
    
    Args:
        campaign_data: Campaign configuration and contact list
        
    Returns:
        Created campaign
    """
    service = CampaignService(db)
    
    try:
        # Resolve company_id from the agent entity.
        # The agent is bound to the tenant company, so this ensures
        # the campaign uses the tenant's phone number, not the app admin's.
        agent_result = await db.execute(
            select(HierarchicalEntity).where(
                HierarchicalEntity.id == campaign_data.agent_id
            )
        )
        agent_entity = agent_result.scalar_one_or_none()
        if not agent_entity:
            raise HTTPException(status_code=404, detail="Agent not found")

        # Use agent's company (tenant) instead of current_user's company (may be APP)
        resolved_company_id = UUID(str(agent_entity.company_id))
        logger.info(
            f"Creating campaign: user company={current_user.company_id}, "
            f"agent company={resolved_company_id} (resolved from agent {campaign_data.agent_id})"
        )

        campaign = await service.create_campaign(
            company_id=resolved_company_id,
            created_by=UUID(str(current_user.id)),
            agent_id=campaign_data.agent_id,
            name=campaign_data.name,
            contact_list=campaign_data.contact_list,
            description=campaign_data.description,
            provider=campaign_data.provider,
            call_script_template=campaign_data.call_script_template,
            scheduled_start=campaign_data.scheduled_start,
            scheduled_end=campaign_data.scheduled_end,
            max_concurrent_calls=campaign_data.max_concurrent_calls,
            max_calls_per_hour=campaign_data.max_calls_per_hour,
            metadata=campaign_data.metadata
        )
        
        return {
            "id": str(campaign.id),
            "name": campaign.name,
            "description": campaign.description,
            "status": campaign.status,
            "total_contacts": campaign.total_contacts,
            "calls_initiated": campaign.calls_initiated,
            "calls_completed": campaign.calls_completed,
            "calls_failed": campaign.calls_failed,
            "created_at": campaign.created_at.isoformat()
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating campaign: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/upload-csv")
async def upload_csv(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Parse and validate uploaded CSV file.
    
    Args:
        file: CSV file with contacts
        
    Returns:
        Parsed contacts with validation results
    """
    service = CampaignService(db)
    
    try:
        # Parse CSV
        contacts = await service.parse_csv(file)
        
        # Validate contacts
        validation = await service.validate_contacts(contacts)
        
        return {
            "total": len(contacts),
            "valid": validation["valid"],
            "invalid": validation["invalid"],
            "errors": validation["errors"],
            "contacts": validation["contacts"][:10],  # Return first 10 for preview
            "all_contacts": validation["contacts"]  # Full list for form submission
        }
        
    except Exception as e:
        logger.error(f"Error parsing CSV: {e}")
        raise HTTPException(status_code=400, detail=str(e))


@router.get("")
async def list_campaigns(
    status: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    List campaigns with real-time call status counts.
    """
    from src.ai.campaign_models import CampaignCall
    from sqlalchemy import func

    service = CampaignService(db)
    
    try:
        campaigns = await service.list_campaigns(
            company_id=UUID(str(current_user.company_id)),
            status=status,
            limit=limit,
            offset=offset
        )

        # Fetch real-time call counts for all campaigns in one query
        campaign_ids = [c.id for c in campaigns]
        if campaign_ids:
            counts_result = await db.execute(
                select(
                    CampaignCall.campaign_id,
                    CampaignCall.status,
                    func.count(CampaignCall.id).label("count"),
                )
                .where(CampaignCall.campaign_id.in_(campaign_ids))
                .group_by(CampaignCall.campaign_id, CampaignCall.status)
            )
            # Build a nested dict: {campaign_id: {status: count}}
            counts_map: dict = {}
            for row in counts_result:
                counts_map.setdefault(row.campaign_id, {})[row.status] = row.count
        else:
            counts_map = {}

        result = []
        for c in campaigns:
            sc = counts_map.get(c.id, {})
            completed = sc.get("completed", 0)
            failed = sc.get("failed", 0)
            calling = sc.get("calling", 0)
            pending = sc.get("pending", 0)

            # Derive effective status
            effective_status = c.status
            if c.status == "completed" and pending > 0:
                effective_status = "running"

            result.append({
                "id": str(c.id),
                "name": c.name,
                "description": c.description,
                "status": effective_status,
                "total_contacts": c.total_contacts,
                "calls_initiated": completed + failed + calling,
                "calls_completed": completed,
                "calls_failed": failed,
                "calls_calling": calling,
                "calls_pending": pending,
                "provider": c.provider,
                "created_at": c.created_at.isoformat()
            })

        return {
            "total": len(result),
            "campaigns": result,
        }
        
    except Exception as e:
        logger.error(f"Error listing campaigns: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{campaign_id}")
async def get_campaign(
    campaign_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get campaign details with real-time call status counts and individual call records.
    """
    from src.ai.campaign_models import CampaignCall
    from sqlalchemy import func

    service = CampaignService(db)
    
    try:
        campaign = await service.get_campaign(campaign_id)
        
        if not campaign:
            raise HTTPException(status_code=404, detail="Campaign not found")

        # Real-time call status counts from CampaignCall table
        status_result = await db.execute(
            select(
                CampaignCall.status,
                func.count(CampaignCall.id).label("count"),
            )
            .where(CampaignCall.campaign_id == campaign_id)
            .group_by(CampaignCall.status)
        )
        status_counts = {row.status: row.count for row in status_result}

        # Fetch individual call records
        calls_result = await db.execute(
            select(CampaignCall)
            .where(CampaignCall.campaign_id == campaign_id)
            .order_by(CampaignCall.created_at.asc())
        )
        calls = calls_result.scalars().all()

        # Derive effective campaign status from call states
        total = campaign.total_contacts
        completed = status_counts.get("completed", 0)
        failed = status_counts.get("failed", 0)
        calling = status_counts.get("calling", 0)
        pending = status_counts.get("pending", 0)

        # If campaign is marked "completed" but has pending calls, it was prematurely closed
        effective_status = campaign.status
        if campaign.status == "completed" and pending > 0:
            effective_status = "running"

        return {
            "id": str(campaign.id),
            "name": campaign.name,
            "description": campaign.description,
            "status": effective_status,
            "total_contacts": total,
            "calls_initiated": completed + failed + calling,
            "calls_completed": completed,
            "calls_failed": failed,
            "calls_calling": calling,
            "calls_pending": pending,
            "max_concurrent_calls": campaign.max_concurrent_calls,
            "max_calls_per_hour": campaign.max_calls_per_hour,
            "scheduled_start": campaign.scheduled_start.isoformat() if campaign.scheduled_start else None,
            "scheduled_end": campaign.scheduled_end.isoformat() if campaign.scheduled_end else None,
            "started_at": campaign.started_at.isoformat() if campaign.started_at else None,
            "completed_at": campaign.completed_at.isoformat() if campaign.completed_at else None,
            "provider": campaign.provider,
            "agent_id": str(campaign.agent_id) if campaign.agent_id else None,
            "created_at": campaign.created_at.isoformat(),
            "calls": [
                {
                    "id": str(c.id),
                    "contact_name": c.contact_data.get("name", "Unknown") if c.contact_data else "Unknown",
                    "contact_phone": c.contact_data.get("phone", "") if c.contact_data else "",
                    "contact_company": c.contact_data.get("company", "") if c.contact_data else "",
                    "status": c.status,
                    "outcome": c.outcome,
                    "outcome_notes": c.outcome_notes,
                    "called_at": c.called_at.isoformat() if c.called_at else None,
                    "completed_at": c.completed_at.isoformat() if c.completed_at else None,
                    "duration_seconds": c.duration_seconds,
                    "voice_session_id": str(c.voice_session_id) if c.voice_session_id else None,
                    "call_sid": c.call_sid,
                }
                for c in calls
            ],
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting campaign: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{campaign_id}/status")
async def get_campaign_status(
    campaign_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get real-time campaign status and metrics.
    
    Args:
        campaign_id: Campaign UUID
        
    Returns:
        Real-time metrics
    """
    service = CampaignService(db)
    
    try:
        status = await service.get_campaign_status(campaign_id)
        
        if not status:
            raise HTTPException(status_code=404, detail="Campaign not found")
        
        return status
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting campaign status: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{campaign_id}/active-calls")
async def get_active_calls(
    campaign_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get currently active calls for a campaign.
    
    Args:
        campaign_id: Campaign UUID
        
    Returns:
        List of active calls
    """
    service = CampaignService(db)
    
    try:
        calls = await service.get_active_calls(campaign_id)
        
        return {
            "campaign_id": str(campaign_id),
            "active_calls": calls
        }
        
    except Exception as e:
        logger.error(f"Error getting active calls: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/{campaign_id}/status")
async def update_campaign_status(
    campaign_id: UUID,
    status: str = Query(..., description="New campaign status"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Update campaign status (start, pause, resume, stop).
    
    Args:
        campaign_id: Campaign UUID
        status: New status ('running', 'paused', 'completed')
        
    Returns:
        Updated campaign
    """
    service = CampaignService(db)
    
    try:
        await service.update_campaign_status(campaign_id, status)
        
        # Enqueue background task if starting campaign
        if status == "running":
            from arq import create_pool
            from arq.connections import RedisSettings
            from src.common.config import settings
            from urllib.parse import urlparse
            
            parsed = urlparse(settings.REDIS_URL or "redis://localhost:6379")
            redis_settings = RedisSettings(host=parsed.hostname or "localhost", port=parsed.port or 6379)
            
            redis = await create_pool(redis_settings)
            await redis.enqueue_job(
                'execute_campaign_task',
                str(campaign_id)
            )
            await redis.close()
            
            logger.info(f"Enqueued campaign execution task for {campaign_id}")
        
        return {
            "campaign_id": str(campaign_id),
            "status": status,
            "message": f"Campaign status updated to {status}"
        }
        
    except Exception as e:
        logger.error(f"Error updating campaign status: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{campaign_id}/download")
async def download_campaign_report(
    campaign_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Download campaign call details as an Excel report.

    Generates a .xlsx file with two sheets:
    - Summary: Campaign metadata and aggregate metrics
    - Call Details: One row per call with contact info, status, timing

    Returns:
        StreamingResponse with Excel file
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from src.ai.campaign_models import CampaignCall
    from src.voice.models import VoiceSession, ConversationHistory
    from sqlalchemy import func

    service = CampaignService(db)

    # Max chars per Excel cell to prevent oversized files
    MAX_CELL_CHARS = 32000

    try:
        campaign = await service.get_campaign(campaign_id)
        if not campaign:
            raise HTTPException(status_code=404, detail="Campaign not found")

        # Fetch call status counts
        status_result = await db.execute(
            select(
                CampaignCall.status,
                func.count(CampaignCall.id).label("count"),
            )
            .where(CampaignCall.campaign_id == campaign_id)
            .group_by(CampaignCall.status)
        )
        status_counts = {row.status: row.count for row in status_result}

        # Fetch all call records
        calls_result = await db.execute(
            select(CampaignCall)
            .where(CampaignCall.campaign_id == campaign_id)
            .order_by(CampaignCall.created_at.asc())
        )
        calls = calls_result.scalars().all()

        # ── Batch-load linked VoiceSession data ──────────────────────────
        voice_session_ids = [
            c.voice_session_id for c in calls if c.voice_session_id
        ]
        voice_sessions_map = {}
        if voice_session_ids:
            vs_result = await db.execute(
                select(VoiceSession).where(
                    VoiceSession.id.in_(voice_session_ids)
                )
            )
            for vs in vs_result.scalars().all():
                voice_sessions_map[vs.id] = vs

        # ── Batch-load ConversationHistory for sessions without conversation_log
        sessions_needing_history = [
            sid for sid in voice_session_ids
            if sid in voice_sessions_map and not voice_sessions_map[sid].conversation_log
        ]
        conversation_history_map = {}
        if sessions_needing_history:
            ch_result = await db.execute(
                select(ConversationHistory)
                .where(ConversationHistory.session_id.in_(sessions_needing_history))
                .order_by(ConversationHistory.turn_number.asc())
            )
            for ch in ch_result.scalars().all():
                conversation_history_map.setdefault(ch.session_id, []).append(ch)

        def _get_transcript_text(call) -> str:
            """Build a formatted transcript string for a campaign call."""
            if not call.voice_session_id or call.voice_session_id not in voice_sessions_map:
                return ""

            vs = voice_sessions_map[call.voice_session_id]

            # Prefer conversation_log JSONB (set during cleanup)
            if vs.conversation_log:
                lines = []
                for turn in vs.conversation_log:
                    speaker = (turn.get("speaker") or "unknown").capitalize()
                    content = turn.get("content", "")
                    if content:
                        lines.append(f"{speaker}: {content}")
                return "\n".join(lines)[:MAX_CELL_CHARS]

            # Fallback: ConversationHistory table
            history = conversation_history_map.get(call.voice_session_id, [])
            if history:
                lines = []
                for ch in history:
                    speaker = (ch.speaker or "unknown").capitalize()
                    content = ch.content or ""
                    if content:
                        lines.append(f"{speaker}: {content}")
                return "\n".join(lines)[:MAX_CELL_CHARS]

            return ""

        def _get_summary(call) -> str:
            """Get call summary from linked VoiceSession."""
            if not call.voice_session_id or call.voice_session_id not in voice_sessions_map:
                return ""
            vs = voice_sessions_map[call.voice_session_id]
            ctx = vs.context_state or {}
            return (ctx.get("call_summary") or "")[:MAX_CELL_CHARS]

        def _get_next_action(call) -> str:
            """Get next action from linked VoiceSession."""
            if not call.voice_session_id or call.voice_session_id not in voice_sessions_map:
                return ""
            vs = voice_sessions_map[call.voice_session_id]
            ctx = vs.context_state or {}
            return (ctx.get("next_action") or "")[:MAX_CELL_CHARS]

        # ── Build Excel workbook ──────────────────────────────────────
        wb = Workbook()

        # Styles
        header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="2B3A67", end_color="2B3A67", fill_type="solid")
        title_font = Font(name="Calibri", bold=True, size=14, color="2B3A67")
        label_font = Font(name="Calibri", bold=True, size=11)
        value_font = Font(name="Calibri", size=11)
        thin_border = Border(
            bottom=Side(style="thin", color="CCCCCC")
        )

        # ── Sheet 1: Summary ──────────────────────────────────────────
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.column_dimensions["A"].width = 25
        ws_summary.column_dimensions["B"].width = 40

        ws_summary.append([])
        title_cell = ws_summary.cell(row=1, column=1, value="Campaign Report")
        title_cell.font = Font(name="Calibri", bold=True, size=18, color="2B3A67")

        summary_data = [
            ("", ""),
            ("Campaign Name", campaign.name),
            ("Description", campaign.description or "—"),
            ("Status", campaign.status.upper()),
            ("Provider", "Tata Tele" if campaign.provider == "tata_tele" else "Twilio"),
            ("", ""),
            ("Total Contacts", campaign.total_contacts),
            ("Calls Completed", status_counts.get("completed", 0)),
            ("Calls Failed", status_counts.get("failed", 0)),
            ("Calls Pending", status_counts.get("pending", 0)),
            ("Calls In Progress", status_counts.get("calling", 0)),
            ("", ""),
            ("Created At", campaign.created_at.strftime("%Y-%m-%d %H:%M:%S") if campaign.created_at else "—"),
            ("Started At", campaign.started_at.strftime("%Y-%m-%d %H:%M:%S") if campaign.started_at else "—"),
            ("Completed At", campaign.completed_at.strftime("%Y-%m-%d %H:%M:%S") if campaign.completed_at else "—"),
            ("", ""),
            ("Max Concurrent Calls", campaign.max_concurrent_calls or 5),
            ("Max Calls/Hour", campaign.max_calls_per_hour or "Unlimited"),
        ]

        for row_data in summary_data:
            ws_summary.append(list(row_data))

        # Style summary labels and values
        for row in ws_summary.iter_rows(min_row=3, max_col=2):
            if row[0].value:
                row[0].font = label_font
                row[0].border = thin_border
            if row[1].value:
                row[1].font = value_font
                row[1].border = thin_border

        # ── Sheet 2: Call Details ─────────────────────────────────────
        ws_calls = wb.create_sheet("Call Details")

        call_headers = [
            "#", "Contact Name", "Phone Number", "Company",
            "Status", "Outcome", "Outcome Notes",
            "Called At", "Completed At", "Duration (sec)",
            "Transcription", "Call Summary", "Next Action",
            "Call SID", "Retry Count"
        ]

        # Write header row
        for col_idx, header in enumerate(call_headers, 1):
            cell = ws_calls.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Set column widths
        col_widths = [6, 25, 20, 20, 12, 15, 30, 22, 22, 14, 50, 40, 40, 25, 12]
        for i, width in enumerate(col_widths, 1):
            ws_calls.column_dimensions[ws_calls.cell(row=1, column=i).column_letter].width = width

        # Status color fills
        status_fills = {
            "completed": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
            "failed": PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
            "calling": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
            "pending": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
        }

        # Write call data rows
        for idx, call in enumerate(calls, 1):
            contact = call.contact_data or {}
            row_data = [
                idx,
                contact.get("name", "Unknown"),
                contact.get("phone", ""),
                contact.get("company", ""),
                call.status or "—",
                call.outcome or "—",
                call.outcome_notes or "",
                call.called_at.strftime("%Y-%m-%d %H:%M:%S") if call.called_at else "—",
                call.completed_at.strftime("%Y-%m-%d %H:%M:%S") if call.completed_at else "—",
                call.duration_seconds if call.duration_seconds is not None else "—",
                _get_transcript_text(call),
                _get_summary(call),
                _get_next_action(call),
                call.call_sid or "—",
                call.retry_count or 0,
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws_calls.cell(row=idx + 1, column=col_idx, value=value)
                cell.font = value_font
                cell.border = thin_border
                # Color-code status column
                if col_idx == 5 and call.status in status_fills:
                    cell.fill = status_fills[call.status]
                # Wrap text for transcript, summary, next action columns
                if col_idx in (11, 12, 13):
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Freeze header row
        ws_calls.freeze_panes = "A2"

        # ── Save to buffer and return ─────────────────────────────────
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        safe_name = campaign.name.replace(" ", "_").replace("/", "-")[:50]
        filename = f"Campaign_Report_{safe_name}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating campaign report: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
