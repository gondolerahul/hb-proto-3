"""
xlsx_engine.py — Deterministic XLSX rendering engine for professional spreadsheets.

Provides pre-built styling helpers and themed components so the LLM
calls methods with structured data instead of writing raw openpyxl code.
"""
import os
from typing import Any, Dict, List, Optional, Tuple

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        NamedStyle,
        PatternFill,
        Side,
        numbers,
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

# ── Theme Definitions ────────────────────────────────────────────────────────

THEMES = {
    "midnight_executive": {
        "header_bg": "1E2761",
        "header_font": "FFFFFF",
        "accent": "7C3AED",
        "row_alt": "F0F0FF",
        "text": "1A1A2E",
        "chart_colors": ["7C3AED", "3B82F6", "10B981", "F59E0B", "EF4444"],
    },
    "forest_moss": {
        "header_bg": "2C5F2D",
        "header_font": "FFFFFF",
        "accent": "97BC62",
        "row_alt": "F0F7E8",
        "text": "1B1B1B",
        "chart_colors": ["97BC62", "4A7C4B", "8FBC8F", "556B2F", "228B22"],
    },
    "coral_energy": {
        "header_bg": "2F3C7E",
        "header_font": "FFFFFF",
        "accent": "F96167",
        "row_alt": "FFF0F0",
        "text": "2F3C7E",
        "chart_colors": ["F96167", "F9E795", "FCB69F", "FF6B6B", "FFA07A"],
    },
    "charcoal_minimal": {
        "header_bg": "36454F",
        "header_font": "FFFFFF",
        "accent": "64B5F6",
        "row_alt": "F5F5F5",
        "text": "212121",
        "chart_colors": ["64B5F6", "5F6B7C", "87919E", "B0B8C1", "26C6DA"],
    },
}


class XlsxEngine:
    """Deterministic XLSX rendering engine with professional styling."""

    def __init__(self, theme: str = "midnight_executive"):
        if openpyxl is None:
            raise RuntimeError("openpyxl is not installed. pip install openpyxl")
        self.wb = Workbook()
        # Remove the default sheet — we'll create named sheets
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
        self.theme_name = theme
        self.t = THEMES.get(theme, THEMES["midnight_executive"])

        # Pre-build styles
        self._header_fill = PatternFill(
            start_color=self.t["header_bg"],
            end_color=self.t["header_bg"],
            fill_type="solid",
        )
        self._header_font = Font(
            name="Calibri", bold=True, color=self.t["header_font"], size=11
        )
        self._alt_fill = PatternFill(
            start_color=self.t["row_alt"],
            end_color=self.t["row_alt"],
            fill_type="solid",
        )
        self._body_font = Font(name="Calibri", size=11, color=self.t["text"])
        self._thin_border = Border(
            bottom=Side(style="thin", color="E0E0E0"),
        )
        self._header_border = Border(
            bottom=Side(style="medium", color=self.t["accent"]),
        )
        self._center = Alignment(horizontal="center", vertical="center")
        self._left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── Sheet Creation ────────────────────────────────────────────────────

    def add_sheet(
        self,
        name: str,
        columns: List[Dict[str, Any]],
        data: List[List[Any]],
        freeze_header: bool = True,
    ):
        """Create a fully styled sheet with headers, data, and auto-width.

        Args:
            name: Sheet tab name
            columns: [{"header": "Name", "type": "text|number|currency|pct|date", "width": 20}]
            data: [[val1, val2, ...], ...]
            freeze_header: Whether to freeze the top row
        Returns:
            The created worksheet
        """
        ws = self.wb.create_sheet(title=name)

        # Write headers
        for ci, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=ci, value=col["header"])
            cell.fill = self._header_fill
            cell.font = self._header_font
            cell.alignment = self._center
            cell.border = self._header_border
            ws.column_dimensions[get_column_letter(ci)].width = col.get("width", 15)

        # Write data rows with banded styling
        for ri, row in enumerate(data, 2):
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = self._body_font
                cell.border = self._thin_border
                cell.alignment = self._left

                # Apply number format based on column type
                if ci <= len(columns):
                    col_type = columns[ci - 1].get("type", "text")
                    if col_type == "currency":
                        cell.number_format = '$#,##0.00'
                    elif col_type == "pct":
                        cell.number_format = '0.0%'
                    elif col_type == "number":
                        cell.number_format = '#,##0'
                    elif col_type == "date":
                        cell.number_format = 'YYYY-MM-DD'

                # Banded row shading
                if ri % 2 == 0:
                    cell.fill = self._alt_fill

        # Freeze header row
        if freeze_header:
            ws.freeze_panes = "A2"

        return ws

    # ── Theming ───────────────────────────────────────────────────────────

    def apply_theme(self, ws):
        """Apply full theme styling to an existing worksheet.

        Assumes row 1 is headers. Applies header colors, banded rows,
        borders, and auto-column-width.
        """
        max_col = ws.max_column
        max_row = ws.max_row

        # Style headers
        for ci in range(1, max_col + 1):
            cell = ws.cell(row=1, column=ci)
            cell.fill = self._header_fill
            cell.font = self._header_font
            cell.alignment = self._center
            cell.border = self._header_border

        # Style data rows
        for ri in range(2, max_row + 1):
            for ci in range(1, max_col + 1):
                cell = ws.cell(row=ri, column=ci)
                cell.font = self._body_font
                cell.border = self._thin_border
                if ri % 2 == 0:
                    cell.fill = self._alt_fill

        # Auto-width columns
        for ci in range(1, max_col + 1):
            max_len = 0
            col_letter = get_column_letter(ci)
            for ri in range(1, max_row + 1):
                val = ws.cell(row=ri, column=ci).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # ── KPI Cards ─────────────────────────────────────────────────────────

    def add_kpi_card(
        self, ws, cell_ref: str, label: str, value: str, format_str: str = ""
    ):
        """Add a KPI card with large value + small label using merged cells.

        Args:
            ws: Target worksheet
            cell_ref: Top-left cell (e.g. "B2")
            label: KPI label (e.g. "Total Revenue")
            value: KPI value (e.g. "$4.2M")
            format_str: Optional number format
        """
        from openpyxl.utils import column_index_from_string

        col_letter = "".join(c for c in cell_ref if c.isalpha())
        row_num = int("".join(c for c in cell_ref if c.isdigit()))
        col_num = column_index_from_string(col_letter)

        # Merge 2 cols × 3 rows for the card
        end_col = get_column_letter(col_num + 1)
        ws.merge_cells(f"{cell_ref}:{end_col}{row_num + 2}")

        # Value cell (large, bold)
        val_cell = ws.cell(row=row_num, column=col_num, value=value)
        val_cell.font = Font(
            name="Calibri", size=28, bold=True, color=self.t["header_bg"]
        )
        val_cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        if format_str:
            val_cell.number_format = format_str

        # Add accent top border
        accent_fill = PatternFill(
            start_color=self.t["accent"],
            end_color=self.t["accent"],
            fill_type="solid",
        )
        for c in range(col_num, col_num + 2):
            border_cell = ws.cell(row=row_num, column=c)
            border_cell.border = Border(
                top=Side(style="thick", color=self.t["accent"])
            )

        # Label below the merged area
        label_cell = ws.cell(row=row_num + 3, column=col_num, value=label)
        label_cell.font = Font(name="Calibri", size=10, color="888888")
        label_cell.alignment = Alignment(horizontal="center")

    # ── Native Charts ─────────────────────────────────────────────────────

    def add_native_chart(
        self,
        ws,
        chart_type: str,
        data_range: str,
        title: str,
        position: str = "E2",
        width: int = 15,
        height: int = 10,
    ):
        """Add a native Excel chart with theme colors.

        Args:
            ws: Target worksheet
            chart_type: "bar", "line", "pie"
            data_range: Data range string e.g. "A1:D5"
            title: Chart title
            position: Cell for chart placement (e.g. "E2")
            width: Chart width in column units
            height: Chart height in row units
        """
        # Parse data range
        parts = data_range.replace(":", "").split()
        if len(parts) < 2:
            # Handle "A1:D5" format
            rng = data_range.split(":")
            if len(rng) == 2:
                start_cell = rng[0]
                end_cell = rng[1]
            else:
                return

        # Create appropriate chart
        if chart_type == "bar":
            chart = BarChart()
            chart.style = 10
        elif chart_type == "line":
            chart = LineChart()
            chart.style = 10
        elif chart_type == "pie":
            chart = PieChart()
        else:
            chart = BarChart()
            chart.style = 10

        chart.title = title
        chart.width = width
        chart.height = height

        # Parse range for data reference
        from openpyxl.utils import column_index_from_string

        start_col_letter = "".join(c for c in start_cell if c.isalpha())
        start_row = int("".join(c for c in start_cell if c.isdigit()))
        end_col_letter = "".join(c for c in end_cell if c.isalpha())
        end_row = int("".join(c for c in end_cell if c.isdigit()))

        start_col = column_index_from_string(start_col_letter)
        end_col = column_index_from_string(end_col_letter)

        # Categories from first column, data from remaining columns
        cats = Reference(ws, min_col=start_col, min_row=start_row + 1, max_row=end_row)
        for col_idx in range(start_col + 1, end_col + 1):
            data_ref = Reference(
                ws, min_col=col_idx, min_row=start_row, max_row=end_row
            )
            chart.add_data(data_ref, titles_from_data=True)

        chart.set_categories(cats)

        # Apply theme colors to series
        colors = self.t["chart_colors"]
        for i, series in enumerate(chart.series):
            color_hex = colors[i % len(colors)]
            series.graphicalProperties.solidFill = color_hex

        ws.add_chart(chart, position)

    # ── Number Formatting ─────────────────────────────────────────────────

    def format_currency(self, ws, cell_range: str):
        """Apply currency formatting ($#,##0.00) to a range."""
        self._apply_format(ws, cell_range, '$#,##0.00')

    def format_percent(self, ws, cell_range: str):
        """Apply percentage formatting (0.0%) to a range."""
        self._apply_format(ws, cell_range, '0.0%')

    def format_number(self, ws, cell_range: str):
        """Apply number formatting (#,##0) to a range."""
        self._apply_format(ws, cell_range, '#,##0')

    def _apply_format(self, ws, cell_range: str, fmt: str):
        """Apply a number format to a cell range like 'C2:C100'."""
        for row in ws[cell_range]:
            for cell in row if isinstance(row, tuple) else [row]:
                cell.number_format = fmt

    # ── Conditional Formatting ────────────────────────────────────────────

    def add_conditional_format(self, ws, cell_range: str, rule_type: str = "color_scale"):
        """Add conditional formatting to a range.

        Args:
            ws: Target worksheet
            cell_range: e.g. "D2:D20"
            rule_type: "color_scale" or "data_bar"
        """
        if rule_type == "color_scale":
            rule = ColorScaleRule(
                start_type="min",
                start_color="F8696B",  # Red
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",  # Yellow
                end_type="max",
                end_color="63BE7B",  # Green
            )
            ws.conditional_formatting.add(cell_range, rule)
        elif rule_type == "data_bar":
            rule = DataBarRule(
                start_type="min",
                end_type="max",
                color=self.t["accent"],
                showValue=True,
            )
            ws.conditional_formatting.add(cell_range, rule)

    # ── Dashboard Setup ───────────────────────────────────────────────────

    def setup_dashboard(self, ws, print_area: str = None):
        """Configure a sheet as a dashboard: freeze panes, hide gridlines, set print area."""
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        if print_area:
            ws.print_area = print_area
        # Set zoom
        ws.sheet_view.zoomScale = 100

    # ── Save ──────────────────────────────────────────────────────────────

    def save(self, output_path: str) -> str:
        """Save the workbook to disk."""
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        self.wb.save(output_path)
        return output_path
