"""Excel and CSV tool for AI Entities — Read, Create, and Update spreadsheets."""

import csv
import json
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from src.ai.tools.base import Tool

try:
    import openpyxl
except ImportError:
    openpyxl = None

BASE_DIR = Path(__file__).resolve().parents[3] / "artifact" / "system-generated"


class ExcelTool(Tool):
    """Tool for reading, creating, and updating Excel (.xlsx) and CSV files.

    Actions:
        read_rows    — Read rows from an existing file
        get_columns  — Get column names from a file
        list_sheets  — List all sheet names in a workbook
        create       — Build a new xlsx workbook with headers and data
        update       — Add/modify rows, add sheets, update cells
    """

    name = "excel_tool"
    description = (
        "Read, create, or update Excel (.xlsx) or CSV files. "
        "Input: JSON with 'action' ('read_rows', 'get_columns', 'list_sheets', "
        "'create', or 'update') and action-specific parameters."
    )

    async def run(self, input_data: str) -> str:
        return await self.run_with_context(input_data, context=None)

    async def run_with_context(self, input_data: str, context=None) -> str:
        try:
            params = json.loads(input_data) if isinstance(input_data, str) else input_data
            action = params.get("action")

            # Actions that don't require an existing file
            if action == "create":
                return await self._create(params, context)

            file_path = params.get("file_path")
            if not file_path:
                return json.dumps({"error": "Missing 'file_path'"})

            if not os.path.exists(file_path):
                fallback_path = os.path.join("uploads", os.path.basename(file_path))
                if os.path.exists(fallback_path):
                    file_path = fallback_path
                else:
                    return json.dumps({"error": f"File not found: {file_path}"})

            if action == "read_rows":
                return self._read_rows(file_path, params)
            elif action == "get_columns":
                return self._get_columns(file_path, params)
            elif action == "list_sheets":
                return self._list_sheets(file_path)
            elif action == "update":
                return await self._update(file_path, params, context)
            else:
                return json.dumps({"error": f"Unknown action: {action}"})

        except Exception as e:
            return json.dumps({"error": f"Excel Tool Error: {str(e)}"})

    # ------------------------------------------------------------------
    # READ (existing)
    # ------------------------------------------------------------------

    def _read_rows(self, file_path: str, params: Dict[str, Any]) -> str:
        limit = params.get("limit", 100)
        if file_path.endswith('.csv'):
            return self._read_csv(file_path, limit)
        else:
            return self._read_excel(file_path, params.get("sheet_name"), limit)

    def _get_columns(self, file_path: str, params: Dict[str, Any]) -> str:
        if file_path.endswith('.csv'):
            with open(file_path, mode='r', encoding='utf-8') as f:
                reader = csv.reader(f)
                headers = next(reader, [])
                return json.dumps({"status": "success", "columns": headers})
        else:
            if not openpyxl:
                return json.dumps({"error": "openpyxl not installed"})
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]
            return json.dumps({"status": "success", "columns": headers})

    def _list_sheets(self, file_path: str) -> str:
        if not openpyxl:
            return json.dumps({"error": "openpyxl not installed"})
        wb = openpyxl.load_workbook(file_path, data_only=True)
        return json.dumps({"status": "success", "sheets": wb.sheetnames})

    def _read_csv(self, file_path: str, limit: int) -> str:
        rows = []
        with open(file_path, mode='r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for i, row in enumerate(reader):
                if i >= limit:
                    break
                rows.append(row)
        return json.dumps({"status": "success", "data": rows, "count": len(rows)})

    def _read_excel(self, file_path: str, sheet_name: Optional[str], limit: int) -> str:
        if not openpyxl:
            return json.dumps({"error": "openpyxl not installed"})

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb[sheet_name] if sheet_name else wb.active

            headers = [cell.value for cell in sheet[1]]
            rows = []

            for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=limit + 1, values_only=True)):
                row_dict = dict(zip(headers, row))
                rows.append(row_dict)

            return json.dumps({"status": "success", "data": rows, "count": len(rows), "sheet_name": sheet.title})
        except Exception as e:
            return json.dumps({"error": f"Failed to read Excel: {str(e)}"})

    # ------------------------------------------------------------------
    # CREATE
    # ------------------------------------------------------------------

    async def _create(self, params: Dict[str, Any], context: Optional[Dict[str, Any]]) -> str:
        if not openpyxl:
            return json.dumps({"error": "openpyxl not installed"})

        wb = openpyxl.Workbook()
        ws = wb.active

        filename = params.get("filename", "spreadsheet.xlsx")
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"

        sheet_name = params.get("sheet_name", "Sheet1")
        ws.title = sheet_name

        # Write headers
        headers = params.get("headers", [])
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(header))
            cell.font = openpyxl.styles.Font(bold=True)

        # Write data rows
        data_rows = params.get("data", [])
        for row_idx, row_data in enumerate(data_rows, 2):
            if isinstance(row_data, dict):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
            elif isinstance(row_data, list):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # Additional sheets
        extra_sheets = params.get("extra_sheets", [])
        for sheet_data in extra_sheets:
            es = wb.create_sheet(title=sheet_data.get("name", "Sheet"))
            s_headers = sheet_data.get("headers", [])
            for col_idx, h in enumerate(s_headers, 1):
                cell = es.cell(row=1, column=col_idx, value=str(h))
                cell.font = openpyxl.styles.Font(bold=True)
            for row_idx, row_data in enumerate(sheet_data.get("data", []), 2):
                if isinstance(row_data, list):
                    for col_idx, value in enumerate(row_data, 1):
                        es.cell(row=row_idx, column=col_idx, value=value)

        # Auto-size columns (approximation)
        for ws_item in wb.worksheets:
            for column_cells in ws_item.columns:
                max_length = 0
                col_letter = column_cells[0].column_letter
                for cell in column_cells:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws_item.column_dimensions[col_letter].width = min(max_length + 2, 50)

        # Save
        company_id = (context or {}).get("company_id", params.get("company_id", "default"))
        date_str = datetime.utcnow().strftime("%Y-%m-%d")
        output_dir = BASE_DIR / str(company_id) / date_str
        output_dir.mkdir(parents=True, exist_ok=True)

        file_path = output_dir / filename
        wb.save(str(file_path))

        # Register artifact
        artifact_id = await self._register_artifact(file_path, filename, company_id, params)

        result = {
            "status": "success",
            "file_path": str(file_path),
            "message": f"Created {filename} with {len(headers)} columns and {len(data_rows)} rows",
        }
        if artifact_id:
            result["artifact_id"] = artifact_id
            result["download_url"] = f"/api/v1/artifacts/{artifact_id}/download"
            result["document_path"] = f"/api/v1/artifacts/{artifact_id}/download"

        return json.dumps(result)

    # ------------------------------------------------------------------
    # UPDATE
    # ------------------------------------------------------------------

    async def _update(self, file_path: str, params: Dict[str, Any], context: Optional[Dict[str, Any]]) -> str:
        if not openpyxl:
            return json.dumps({"error": "openpyxl not installed"})

        wb = openpyxl.load_workbook(file_path)
        changes_made = 0

        sheet_name = params.get("sheet_name")
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        # Append rows
        append_rows = params.get("append_rows", [])
        for row_data in append_rows:
            if isinstance(row_data, list):
                ws.append(row_data)
            elif isinstance(row_data, dict):
                headers = [cell.value for cell in ws[1]]
                row = [row_data.get(h, "") for h in headers]
                ws.append(row)
            changes_made += 1

        # Update specific cells
        cell_updates = params.get("update_cells", [])
        for update in cell_updates:
            row = update.get("row")
            col = update.get("column")
            value = update.get("value")
            if row and col:
                ws.cell(row=row, column=col, value=value)
                changes_made += 1

        # Add new sheet
        new_sheet = params.get("add_sheet")
        if new_sheet:
            name = new_sheet.get("name", "NewSheet")
            ns = wb.create_sheet(title=name)
            headers = new_sheet.get("headers", [])
            for col_idx, h in enumerate(headers, 1):
                cell = ns.cell(row=1, column=col_idx, value=str(h))
                cell.font = openpyxl.styles.Font(bold=True)
            for row_idx, row_data in enumerate(new_sheet.get("data", []), 2):
                if isinstance(row_data, list):
                    for col_idx, value in enumerate(row_data, 1):
                        ns.cell(row=row_idx, column=col_idx, value=value)
            changes_made += 1

        save_path = params.get("save_as", file_path)
        wb.save(save_path)

        return json.dumps({
            "status": "success",
            "file_path": save_path,
            "changes_made": changes_made,
            "message": f"Updated spreadsheet with {changes_made} changes",
        })

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    async def _register_artifact(
        file_path: Path, filename: str, company_id: str, params: Dict[str, Any]
    ) -> str | None:
        """Register file as an artifact. Returns artifact ID or None."""
        if company_id and company_id != "default":
            try:
                from uuid import UUID as _UUID
                from src.common.database import AsyncSessionLocal
                from src.ai.artifact_service import ArtifactService, ORIGIN_SYSTEM

                async with AsyncSessionLocal() as _db:
                    art_svc = ArtifactService(_db)
                    with open(file_path, "rb") as f:
                        file_bytes = f.read()
                    artifact = await art_svc.save_artifact(
                        file_bytes=file_bytes,
                        file_name=filename,
                        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        file_category="documents",
                        origin=ORIGIN_SYSTEM,
                        company_id=_UUID(str(company_id)),
                        purpose=params.get("purpose", "AI-generated spreadsheet"),
                        generated_by=params.get("generated_by", "excel_tool"),
                    )
                    return str(artifact.id)
            except Exception:
                pass
        return None

    def get_function_schema(self) -> Dict[str, Any]:
        return {
            "name": self.name,
            "description": self.description,
            "parameters": {
                "type": "object",
                "properties": {
                    "action": {
                        "type": "string",
                        "enum": ["read_rows", "get_columns", "list_sheets", "create", "update"],
                        "description": "The action to perform",
                    },
                    "file_path": {
                        "type": "string",
                        "description": "Path to existing Excel/CSV file (for read/update)",
                    },
                    "filename": {
                        "type": "string",
                        "description": "Output filename for 'create' (e.g. report.xlsx)",
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Sheet name (for read/update/create)",
                    },
                    "headers": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Column headers for 'create'",
                    },
                    "data": {
                        "type": "array",
                        "description": "Data rows for 'create' — array of arrays or objects",
                        "items": {"type": "object"},
                    },
                    "extra_sheets": {
                        "type": "array",
                        "description": "Additional sheets for 'create': [{name, headers, data}]",
                        "items": {"type": "object"},
                    },
                    "append_rows": {
                        "type": "array",
                        "description": "Rows to append for 'update' — arrays or objects",
                        "items": {"type": "object"},
                    },
                    "update_cells": {
                        "type": "array",
                        "description": "Cells to update: [{row, column, value}]",
                        "items": {"type": "object"},
                    },
                    "add_sheet": {
                        "type": "object",
                        "description": "New sheet to add: {name, headers, data}",
                    },
                    "save_as": {
                        "type": "string",
                        "description": "Optional save-as path for 'update'",
                    },
                    "limit": {
                        "type": "integer",
                        "description": "Max rows to read (default: 100)",
                    },
                },
                "required": ["action"],
            },
        }
