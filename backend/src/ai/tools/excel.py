"""Excel and CSV reading tool for AI Entities."""

import csv
import json
import os
from typing import Any, Dict, List, Optional
from src.ai.tools.base import Tool

try:
    import openpyxl
except ImportError:
    openpyxl = None

class ExcelTool(Tool):
    """Tool for reading data from Excel (.xlsx, .xls) and CSV files.
    
    Actions:
    1. read_rows: Read rows from a file.
    2. get_columns: Get column names from a file.
    """
    
    name = "excel_tool"
    description = (
        "Read data from Excel (.xlsx, .xls) or CSV files. "
        "Input should be a JSON object with 'action' ('read_rows' or 'get_columns') "
        "and 'file_path'. Optional: 'sheet_name', 'limit'."
    )

    async def run(self, input_data: str) -> str:
        try:
            params = json.loads(input_data)
            action = params.get("action")
            file_path = params.get("file_path")
            
            if not file_path:
                return json.dumps({"error": "Missing 'file_path'"})
            
            if not os.path.exists(file_path):
                # Check in uploads directory as a fallback
                fallback_path = os.path.join("uploads", os.path.basename(file_path))
                if os.path.exists(fallback_path):
                    file_path = fallback_path
                else:
                    return json.dumps({"error": f"File not found: {file_path}"})

            if action == "read_rows":
                return self._read_rows(file_path, params)
            elif action == "get_columns":
                return self._get_columns(file_path, params)
            else:
                return json.dumps({"error": f"Unknown action: {action}"})

        except Exception as e:
            return json.dumps({"error": f"Excel Tool Error: {str(e)}"})

    def _read_rows(self, file_path: str, params: Dict[str, Any]) -> str:
        limit = params.get("limit", 100)
        
        if file_path.endswith('.csv'):
            return self._read_csv(file_path, limit)
        else:
            return self._read_excel(file_path, params.get("sheet_name"), limit)

    def _get_columns(self, file_path: str, params: Dict[str, Any]) -> str:
        if file_path.endswith('.csv'):
            with open(file_path, mode='r', encoding='utf-8') as f:
                reader = csv.reader(f)
                headers = next(reader, [])
                return json.dumps({"status": "success", "columns": headers})
        else:
            if not openpyxl:
                return json.dumps({"error": "openpyxl not installed"})
            wb = openpyxl.load_all(file_path) if hasattr(openpyxl, 'load_all') else openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]
            return json.dumps({"status": "success", "columns": headers})

    def _read_csv(self, file_path: str, limit: int) -> str:
        rows = []
        with open(file_path, mode='r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for i, row in enumerate(reader):
                if i >= limit:
                    break
                rows.append(row)
        return json.dumps({"status": "success", "data": rows, "count": len(rows)})

    def _read_excel(self, file_path: str, sheet_name: Optional[str], limit: int) -> str:
        if not openpyxl:
            return json.dumps({"error": "openpyxl not installed"})
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb[sheet_name] if sheet_name else wb.active
            
            headers = [cell.value for cell in sheet[1]]
            rows = []
            
            for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=limit+1, values_only=True)):
                row_dict = dict(zip(headers, row))
                rows.append(row_dict)
                
            return json.dumps({"status": "success", "data": rows, "count": len(rows), "sheet_name": sheet.title})
        except Exception as e:
            return json.dumps({"error": f"Failed to read Excel: {str(e)}"})

    def get_function_schema(self) -> Dict[str, Any]:
        return {
            "name": self.name,
            "description": self.description,
            "parameters": {
                "type": "object",
                "properties": {
                    "action": {
                        "type": "string",
                        "enum": ["read_rows", "get_columns"],
                        "description": "The action to perform"
                    },
                    "file_path": {
                        "type": "string",
                        "description": "Path to the Excel or CSV file"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Optional: Name of the sheet to read (Excel only)"
                    },
                    "limit": {
                        "type": "integer",
                        "description": "Optional: Max number of rows to read"
                    }
                },
                "required": ["action", "file_path"]
            }
        }
